"""Document processing with OCR and text extraction"""
import os
import logging
from typing import Dict, Optional
from pathlib import Path
import PyPDF2
import pytesseract
from PIL import Image
import openpyxl
import pandas as pd
import io

logger = logging.getLogger(__name__)

class DocumentProcessor:
    """Process various document types and extract text"""
    
    @staticmethod
    async def process_document(file_path: str, file_type: str) -> Dict:
        """
        Process document and extract text based on file type
        
        Args:
            file_path: Path to uploaded file
            file_type: MIME type or file extension
            
        Returns:
            Dict with extracted text and metadata
        """
        try:
            if 'pdf' in file_type.lower():
                return await DocumentProcessor._process_pdf(file_path)
            elif any(img_type in file_type.lower() for img_type in ['jpg', 'jpeg', 'png', 'image']):
                return await DocumentProcessor._process_image(file_path)
            elif 'xlsx' in file_type.lower() or 'excel' in file_type.lower():
                return await DocumentProcessor._process_excel(file_path)
            elif 'csv' in file_type.lower():
                return await DocumentProcessor._process_csv(file_path)
            elif 'text' in file_type.lower() or 'txt' in file_type.lower():
                return await DocumentProcessor._process_text(file_path)
            else:
                return {
                    'success': False,
                    'error': f'Unsupported file type: {file_type}',
                    'extracted_text': ''
                }
        except Exception as e:
            logger.error(f'Document processing error: {str(e)}')
            return {
                'success': False,
                'error': str(e),
                'extracted_text': ''
            }
    
    @staticmethod
    async def _process_pdf(file_path: str) -> Dict:
        """Extract text from PDF"""
        try:
            with open(file_path, 'rb') as file:
                reader = PyPDF2.PdfReader(file)
                text = ''
                
                # Extract metadata
                metadata = {
                    'pages': len(reader.pages),
                    'title': reader.metadata.title if reader.metadata else None
                }
                
                # Extract text from all pages
                for page_num, page in enumerate(reader.pages):
                    page_text = page.extract_text()
                    if page_text:
                        text += f'\\n--- Page {page_num + 1} ---\\n{page_text}'
                
                return {
                    'success': True,
                    'extracted_text': text.strip(),
                    'metadata': metadata,
                    'extraction_method': 'PyPDF2'
                }
        except Exception as e:
            logger.error(f'PDF processing error: {str(e)}')
            return {
                'success': False,
                'error': str(e),
                'extracted_text': ''
            }
    
    @staticmethod
    async def _process_image(file_path: str) -> Dict:
        """Extract text from image using OCR"""
        try:
            image = Image.open(file_path)
            
            # Get image metadata
            metadata = {
                'size': image.size,
                'format': image.format,
                'mode': image.mode
            }
            
            # Perform OCR
            text = pytesseract.image_to_string(image)
            
            # Get confidence score
            try:
                ocr_data = pytesseract.image_to_data(image, output_type=pytesseract.Output.DICT)
                confidences = [int(conf) for conf in ocr_data['conf'] if int(conf) > 0]
                avg_confidence = sum(confidences) / len(confidences) if confidences else 0
                metadata['confidence'] = avg_confidence / 100.0
            except:
                metadata['confidence'] = None
            
            return {
                'success': True,
                'extracted_text': text.strip(),
                'metadata': metadata,
                'extraction_method': 'Tesseract OCR'
            }
        except Exception as e:
            logger.error(f'Image processing error: {str(e)}')
            return {
                'success': False,
                'error': str(e),
                'extracted_text': ''
            }
    
    @staticmethod
    async def _process_excel(file_path: str) -> Dict:
        """Extract text from Excel file"""
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            
            text_parts = []
            metadata = {
                'sheets': workbook.sheetnames,
                'sheet_count': len(workbook.sheetnames)
            }
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text_parts.append(f'\\n=== Sheet: {sheet_name} ===\\n')
                
                for row in sheet.iter_rows(values_only=True):
                    row_text = ' | '.join(str(cell) if cell is not None else '' for cell in row)
                    if row_text.strip():
                        text_parts.append(row_text)
            
            return {
                'success': True,
                'extracted_text': '\\n'.join(text_parts),
                'metadata': metadata,
                'extraction_method': 'openpyxl'
            }
        except Exception as e:
            logger.error(f'Excel processing error: {str(e)}')
            return {
                'success': False,
                'error': str(e),
                'extracted_text': ''
            }
    
    @staticmethod
    async def _process_csv(file_path: str) -> Dict:
        """Extract text from CSV file"""
        try:
            df = pd.read_csv(file_path)
            
            metadata = {
                'rows': len(df),
                'columns': list(df.columns)
            }
            
            # Convert to text representation
            text = df.to_string(index=False)
            
            return {
                'success': True,
                'extracted_text': text,
                'metadata': metadata,
                'extraction_method': 'pandas'
            }
        except Exception as e:
            logger.error(f'CSV processing error: {str(e)}')
            return {
                'success': False,
                'error': str(e),
                'extracted_text': ''
            }
    
    @staticmethod
    async def _process_text(file_path: str) -> Dict:
        """Extract text from plain text file"""
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                text = file.read()
            
            return {
                'success': True,
                'extracted_text': text,
                'metadata': {'encoding': 'utf-8'},
                'extraction_method': 'direct'
            }
        except UnicodeDecodeError:
            # Try with different encoding
            try:
                with open(file_path, 'r', encoding='latin-1') as file:
                    text = file.read()
                
                return {
                    'success': True,
                    'extracted_text': text,
                    'metadata': {'encoding': 'latin-1'},
                    'extraction_method': 'direct'
                }
            except Exception as e:
                logger.error(f'Text processing error: {str(e)}')
                return {
                    'success': False,
                    'error': str(e),
                    'extracted_text': ''
                }
